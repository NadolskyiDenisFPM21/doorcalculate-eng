import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import io
from django.conf import settings

from html2excel import ExcelParser


def create(path):
    input_file = str(settings.BASE_DIR) + path
    parser = ExcelParser(input_file)
    parser.to_excel(str(settings.BASE_DIR / 'polls/document_gen/to_excel.xlsx'))  
    wb = openpyxl.load_workbook(str(settings.BASE_DIR / 'polls/document_gen/to_excel.xlsx'))
    sheet = wb.active

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000')
            )
    with io.BytesIO() as buffer:
        wb.save(buffer)
        return buffer.getvalue()
